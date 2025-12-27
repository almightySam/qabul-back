from django.shortcuts import render, redirect
from .models import Education, Till, UserModel, QabulHolati
from myapp.forms import UserRegistrationForm 
from django.http import HttpResponse
import csv
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook, load_workbook
from django.conf import settings
import os
import logging
from django.core.paginator import Paginator
from datetime import datetime
from phonenumber_field.phonenumber import PhoneNumber
from phonenumbers import PhoneNumber
from threading import Lock
import phonenumbers
from phonenumbers import PhoneNumberFormat
import pandas as pd
from openpyxl.styles import Font, Alignment

# Logger sozlamasi
logger = logging.getLogger(__name__)

# Excel faylga sinxron yozish uchun lock
excel_lock = Lock()

def save_to_excel(file_path, data):
    try:
        with excel_lock:
            headers = [
                "F.I.SH", "Telefon raqami", "Yashash manzili", 
                "Ta'lim", "Fan", "Til bilish darajasi", 
                "Shahodatnoma", "ID karta", "Pasport", "Rasm", "Roâ€˜yxatdan oâ€˜tish sanasi"
            ]
            
            if len(data) != len(headers):
                logger.error(f"Data uzunligi sarlavhalar bilan mos kelmadi: {len(data)} vs {len(headers)}")
                raise ValueError(f"Ma'lumotlar soni sarlavhalar bilan mos kelmadi: {len(data)} berildi, {len(headers)} kutilmoqda.")

            new_data = pd.DataFrame([data], columns=headers)
            
            if os.path.exists(file_path):
                try:
                    existing_data = pd.read_excel(file_path, engine='openpyxl')
                    if list(existing_data.columns) != headers:
                        logger.warning(f"Fayldagi sarlavhalar mos kelmadi: {existing_data.columns}. Sarlavhalar moslashtirilmoqda.")
                        existing_data = existing_data.reindex(columns=headers, fill_value='')
                    updated_data = pd.concat([existing_data, new_data], ignore_index=True)
                except Exception as e:
                    logger.error(f"Mavjud faylni o'qishda xato: {type(e).__name__}: {str(e)}")
                    updated_data = new_data
            else:
                updated_data = new_data
            
            updated_data.to_excel(file_path, index=False, engine='openpyxl')
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            logger.info(f"Ma'lumotlar {file_path} ga muvaffaqiyatli saqlandi: {data[0]}")
    except Exception as e:
        logger.error(f"Excel faylga yozishda xato: {type(e).__name__}: {str(e)}")
        raise Exception(f"Ma'lumotlarni Excel faylga saqlashda xatolik yuz berdi: {str(e)}. Iltimos, qaytadan urinib ko'ring.")

def index(request):
    # Qabul holatini tekshirish
    qabul_holati = QabulHolati.objects.first()
    
    if qabul_holati and not qabul_holati.ochiq:
        return HttpResponse("""
        <html><body style="text-align:center; padding:50px;">
        <h2>Online Qabul Tugallandi</h2>
        <p>{}</p>
        <a href="/">Bosh sahifa</a>
        </body></html>
        """.format(qabul_holati.xabar))
    
    fans = Education.objects.all()
    tills = Till.objects.all()
    
    if request.method == "POST":
        form = UserRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            excel_file_path = os.path.join(settings.MEDIA_ROOT, 'users_data.xlsx')
            save_to_excel(excel_file_path, [
                user.name, str(user.number), user.shaxar, user.education,
                user.userFan.title, user.userTill.title,
                str(user.files.url) if user.files else '',
                str(user.id_card.url) if user.id_card else '',
                str(user.pasport.url) if user.pasport else '',
                str(user.photo.url) if user.photo else '',
                user.created.strftime('%Y-%m-%d %H:%M') if user.created else ''
            ])
            request.session['success_data'] = {
                "name": user.name,
                "number": str(user.number),
                "shaxar": user.shaxar,
                "education": user.education,
                "userFan": user.userFan.title,
                "userTill": user.userTill.title,
                "files_url": str(user.files.url) if user.files else '',
                "id_card_url": str(user.id_card.url) if user.id_card else '',
                "pasport_url": str(user.pasport.url) if user.pasport else '',
                "photo_url": str(user.photo.url) if user.photo else '',
                "created": user.created.strftime('%Y-%m-%d %H:%M') if user.created else ''
            }
            return redirect('success')
        else:
            return render(request, "signup_form.html", {
                "error": form.errors,
                "fans": fans,
                "tills": tills,
                "form": form
            })
    else:
        form = UserRegistrationForm()
    
    return render(request, "signup_form.html", {'fans': fans, 'tills': tills, 'form': form})

def success(request):
    success_data = request.session.get('success_data', {})
    request.session.pop('success_data', None)
    return render(request, "success.html", {"data": success_data})

@login_required
def table(request):
    users = UserModel.objects.all().order_by('-created')
    
    if 'download' in request.GET:
        response = HttpResponse(content_type='text/csv')
        filename = f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow([
            '#', 'F.I.SH', 'Telefon raqami', 'Yashash manzili', "O‘qish joyi", 
            'Fan', 'Til bilish darajasi', 'Shahodatnoma', 'ID karta', 'Pasport', 'Rasm', 'Yaratilgan vaqt'
        ])
        
        for index, user in enumerate(users, start=1):
            writer.writerow([
                index, user.name, str(user.number), user.shaxar, user.education, 
                user.userFan.title, user.userTill.title, 
                str(user.files.url) if user.files else '', 
                str(user.id_card.url) if user.id_card else '', 
                str(user.pasport.url) if user.pasport else '', 
                str(user.photo.url) if user.photo else '',
                user.created.strftime('%Y-%m-%d %H:%M') if user.created else ''
            ])
        
        return response
    
    paginator = Paginator(users, 500)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'table.html', {
    'users': page_obj,
    'page_obj': page_obj,
})

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect("table")
        else:
            return render(request, "login.html", {"error": "Foydalanuvchi nomi yoki parol noto‘g‘ri!"})
    
    return render(request, "login.html")